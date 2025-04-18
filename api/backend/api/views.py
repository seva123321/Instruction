from datetime import datetime

from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.db import IntegrityError, models
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
import numpy as np
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import SearchFilter
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from api.admin import dashboard_callback
from api.models import (
    User,
    Instruction,
    Tests,
    Video,
    NormativeLegislation,
    Notification,
    InstructionResult,
)
from api.serializers import (
    AdminUserSerializer,
    InstructionSerializer,
    InstructionListSerializer,
    UserProfileSerializer,
    TestSerializer,
    TestListSerializer,
    SignUpSerializer,
    TestResultSerializer,
    VideoSerializer,
    TestResultCreateSerializer,
    NormativeLegislationSerializer,
    InstructionResultSerializer,
    InstructionResultGetSerializer,
    NotificationSerializer,
    RatingSerializer,
)
from api.permissions import IsAdminPermission
from backend.constants import ME


@extend_schema(
    tags=["User"],
    description="Получение, создание, изменение и удаление пользователей.",
)
class UserViewSet(ModelViewSet):
    """Представление для операций с пользователями."""

    queryset = User.objects.all()
    serializer_class = AdminUserSerializer
    permission_classes = (IsAdminPermission,)
    filter_backends = (SearchFilter,)
    search_fields = ("last_name",)
    http_method_names = ("get", "post", "patch", "delete")

    def get_queryset(self):
        """Оптимизация запросов к БД."""
        queryset = super().get_queryset()
        if self.action == "profile":
            return queryset.prefetch_related("badges__badge", "current_rank")
        return queryset

    def get_serializer_class(self):
        """Определяем сериализатор в зависимости от действия."""
        if self.action == "profile":
            return UserProfileSerializer
        return super().get_serializer_class()

    @action(
        detail=False,
        methods=["GET", "PATCH"],
        url_path=ME,
        url_name=ME,
        permission_classes=(IsAuthenticated,),
    )
    def profile(self, request):
        """Представление профиля текущего пользователя."""
        user = request.user

        if request.method == "GET":
            serializer = self.get_serializer(user)
            return Response(serializer.data, status=status.HTTP_200_OK)

        elif request.method == "PATCH":
            serializer = self.get_serializer(
                user,
                data=request.data,
                partial=True,
                context={"request": request},
            )
            serializer.is_valid(raise_exception=True)

            try:
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            except IntegrityError as e:
                return Response(
                    {"error": f"Ошибка сохранения данных. {e}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )


@extend_schema(tags=["Rating"], description="Рейтинг пользователей.")
class RatingViewSet(viewsets.ReadOnlyModelViewSet):
    """Представление для получения рейтинга пользователей."""

    queryset = User.objects.all()
    serializer_class = RatingSerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (SearchFilter,)
    search_fields = ("last_name",)
    http_method_names = ("get",)

    def get_queryset(self):
        """Оптимизация запросов к БД."""
        return (
            super()
            .get_queryset()
            .prefetch_related("badges__badge", "current_rank")
            .order_by("-experience_points")
        )


@extend_schema(tags=["SignUp"], description="Регистрация пользователей.")
class SignUpView(APIView):
    """Представление для регистрации новых пользователей."""

    permission_classes = (AllowAny,)

    def post(self, request):
        """Создает нового пользователя."""
        serializer = SignUpSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            # Используем serializer.save() вместо прямого создания
            user = serializer.save()
        except IntegrityError as e:
            error_messages = {
                "api_user_email": {
                    "email": "Пользователь с таким email уже существует"
                },
                "api_user_mobile_phone": {
                    "mobile_phone": "Пользователь с таким номером телефона уже существует"
                },
                "api_user_face_descriptor": {
                    "face_descriptor": "Такой дескриптор лица уже существует"
                },
            }

            for db_error, message in error_messages.items():
                if db_error in str(e):
                    raise ValidationError(message)

            raise ValidationError(
                {"detail": "Ошибка при создании пользователя"}
            )
        except Exception as e:
            raise ValidationError({"detail": str(e)})

        return Response(
            {"id": user.id, "email": user.email},
            status=status.HTTP_201_CREATED,
        )


@extend_schema(tags=["Login"], description="Аутентификация.")
class LoginView(APIView):
    """Представление для входа через сессии"""

    permission_classes = (AllowAny,)

    def post(self, request, *args, **kwargs):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response(
                {
                    "detail": "Требуется email и пароль",
                    "errors": {
                        "email": "Обязательное поле" if not email else None,
                        "password": (
                            "Обязательное поле" if not password else None
                        ),
                    },
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            user = authenticate(request, username=email, password=password)

            if user is None:
                try:
                    User.objects.get(email=email)
                    error_msg = "Неверный пароль"
                    error_field = "password"
                except User.DoesNotExist:
                    error_msg = "Пользователь с таким email не найден"
                    error_field = "email"

                return Response(
                    {
                        "detail": "Ошибка аутентификации",
                        "errors": {error_field: error_msg},
                    },
                    status=status.HTTP_401_UNAUTHORIZED,
                )
            if user.is_staff:
                return Response(
                    {
                        "detail": "Администраторы могут входить только через /admin/login/",
                        "errors": {"admin": "/admin/login/"},
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )
            login(request, user)

            return Response(
                {
                    "detail": "Успешный вход",
                    "user_id": user.id,
                    "email": user.email,
                    "first_name": user.first_name,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"detail": "Произошла ошибка при входе", "error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@extend_schema(tags=["LoginFace"], description="Аутентификация по лицу.")
class FaceLoginView(APIView):
    """Аутентификация по лицу"""

    permission_classes = (AllowAny,)

    def post(self, request):
        # 1. Получаем дескриптор лица из запроса
        face_descriptor = request.data.get("face_descriptor")
        if not face_descriptor or len(face_descriptor) != 128:
            return Response(
                {
                    "error": "Неправильный формат дескриптора"
                    " - должно быть 128 элементов"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 2. Преобразуем в numpy array
        try:
            input_descriptor = np.array(face_descriptor, dtype=np.float32)
        except Exception as e:
            return Response(
                {"error": f"Неправильный формат дескриптора: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 3. Ищем ближайшего пользователя
        best_match = None
        min_distance = float("inf")

        for user in User.objects.exclude(face_descriptor__isnull=True):
            try:
                # Преобразуем дескриптор из БД
                stored_descriptor = np.array(
                    eval(user.face_descriptor), dtype=np.float32
                )

                # Проверяем размерность дескриптора
                if stored_descriptor.shape != (128,):
                    continue  # Пропускаем некорректные записи

                # Вычисляем евклидово расстояние
                distance = np.linalg.norm(input_descriptor - stored_descriptor)

                # Проверяем пороговое значение
                if (
                    distance < min_distance
                    and distance < settings.FACE_MATCH_THRESHOLD
                ):
                    min_distance = distance
                    best_match = user

            except Exception as e:
                print(
                    f"Error processing user {user.id} face descriptor: {str(e)}"
                )
                continue

        # 4. Проверяем результат
        if best_match:
            if best_match.is_staff:
                return Response(
                    {
                        "detail": "Администраторы могут входить только через /admin/login/",
                        "errors": {"admin": "/admin/login/"},
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )
            login(request, best_match)
            return Response(
                {
                    "detail": "Успешный вход",
                    "user_id": best_match.id,
                    "email": best_match.email,
                    "first_name": best_match.first_name,
                }
            )

        return Response(
            {"error": "Лицо не распознано или пользователь не существует"},
            status=status.HTTP_401_UNAUTHORIZED,
        )


class LogoutView(APIView):
    """Представление для выхода из системы"""
    permission_classes = (AllowAny,)

    permission_classes = (AllowAny,)

    def post(self, request):
        logout(request)
        return Response(
            {"detail": "Успешный выход"}, status=status.HTTP_200_OK
        )


@extend_schema(tags=["Instruction"], description="Получение интруктажей.")
class InstructionViewSet(viewsets.ReadOnlyModelViewSet):
    """Представление для получения инструктажа."""

    queryset = Instruction.objects.all()
    serializer_class = InstructionSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)

        first_instruction = Instruction.objects.first()
        if first_instruction:
            response.data["first_instruction"] = InstructionSerializer(
                first_instruction, context=self.get_serializer_context()
            ).data

        return response

    def get_serializer_class(self):
        """Определяет сериализатор в зависимости от действия."""
        if self.action == "list":
            return InstructionListSerializer
        return InstructionSerializer

    def get_queryset(self):
        user_position = self.request.user.position
        return Instruction.objects.filter(
            models.Q(position=user_position) | models.Q(position__isnull=True)
        )


@extend_schema(tags=["Tests"], description="Получение тестов.")
class TestViewSet(viewsets.ReadOnlyModelViewSet):
    """Представление для получения тестов."""

    queryset = Tests.objects.prefetch_related(
        "questions", "questions__answers", "questions__reference_link"
    ).all()
    serializer_class = TestSerializer
    permission_classes = (IsAuthenticated,)

    def get_serializer_class(self):
        """Определяет сериализатор в зависимости от действия."""
        if self.action == "list":
            return TestListSerializer
        return TestSerializer

    def get_queryset(self):
        user_position = self.request.user.position
        return Tests.objects.filter(
            models.Q(position=user_position) | models.Q(position__isnull=True)
        ).prefetch_related(
            "questions", "questions__answers", "questions__reference_link"
        )


@extend_schema(
    tags=["TestResult"],
    description="Сохранение результатов тестов пользователя.",
)
class TestResultCreateView(APIView):
    """API для сохранения результатов тестирования."""

    permission_classes = (IsAuthenticated,)

    def post(self, request):
        serializer = TestResultCreateSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.validated_data["user"] = request.user
        test_result = serializer.save()

        return Response(
            TestResultSerializer(
                test_result, context={"request": request}
            ).data,
            status=status.HTTP_201_CREATED,
        )


@extend_schema(tags=["NLAs"], description="Получение НПА.")
class NormativeLegislationViewSet(viewsets.ReadOnlyModelViewSet):
    """Представление для получения видео."""

    queryset = NormativeLegislation.objects.all()
    serializer_class = NormativeLegislationSerializer
    permission_classes = (IsAuthenticated,)
    ordering = ("-date",)
    ordering_fields = ("date", "title")


@extend_schema(tags=["Tests"], description="Получение видео.")
class VideoViewSet(viewsets.ReadOnlyModelViewSet):
    """Представление для получения видео."""

    queryset = Video.objects.all()
    serializer_class = VideoSerializer
    permission_classes = (IsAuthenticated,)
    ordering = ("-date",)


@extend_schema(
    tags=["InstructionResult"],
    description="Сохранение результатов прохождения инструктажа пользователя.",
)
class InstructionResultView(APIView):
    """API для сохранения результатов прохождения инструктажа."""

    permission_classes = (IsAuthenticated,)

    def get(self, request):
        """Получение результатов инструктажа для текущего пользователя."""
        instruction_results = InstructionResult.objects.filter(
            user=request.user
        ).order_by("-date")

        serializer = InstructionResultGetSerializer(
            instruction_results, many=True, context={"request": request}
        )
        return Response(serializer.data)

    def post(self, request):
        serializer = InstructionResultSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)

        try:
            instruction_result = serializer.save()
            return Response(
                {
                    "status": "success",
                    "instruction_result_id": instruction_result.id,
                    "is_passed": instruction_result.result,
                },
                status=status.HTTP_201_CREATED,
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_400_BAD_REQUEST
            )


@extend_schema(
    tags=["Notification"],
    description="Получение, чтение уведомлений.",
)
class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)

    @action(detail=True, methods=["post"])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({"status": "marked as read"})

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(
            is_read=True
        )
        return Response({"status": "all marked as read"})


@staff_member_required
def export_to_excel(request):
    # Получаем данные
    context = {}
    dashboard_callback(request, context)

    # Создаем Excel файл
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = (
        f'attachment; filename="full_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Полный отчёт"

    # Стили
    header_fill = PatternFill(
        start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    center_aligned = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 1. Шапка отчёта
    ws.merge_cells("A1:E1")
    ws["A1"] = "Полный отчёт по обучению сотрудников"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center_aligned

    ws["A2"] = "Дата формирования отчёта:"
    ws["B2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["B2"].font = Font(bold=True)

    # 2. Основная статистика
    ws.append([])
    ws.append(["Детальная статистика по тестам и инструктажам"])
    ws.merge_cells("A4:E4")
    ws["A4"].font = Font(size=12, bold=True)
    ws["A4"].fill = PatternFill(
        start_color="70AD47", end_color="70AD47", fill_type="solid"
    )
    ws["A4"].font = Font(color="FFFFFF")
    ws["A4"].alignment = center_aligned

    # Заголовки таблицы
    headers = [
        "Тип проверки",
        "Всего попыток",
        "Успешно пройдено",
        "Не пройдено",
        "Процент успеха",
    ]
    ws.append(headers)

    # Стили для заголовков
    for col in range(1, 6):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    # Данные тестов
    test_total = context["test_stats"]["total"]
    test_passed = context["test_stats"]["passed"]
    test_failed = context["test_stats"]["failed"]
    test_success_rate = (
        (test_passed / test_total * 100) if test_total > 0 else 0
    )

    ws.append(
        [
            "Тесты",
            test_total,
            test_passed,
            test_failed,
            f"{test_success_rate:.1f}%",
        ]
    )

    # Данные инструктажей
    instr_total = context["instruction_stats"]["total"]
    instr_passed = context["instruction_stats"]["passed"]
    instr_failed = context["instruction_stats"]["failed"]
    instr_success_rate = (
        (instr_passed / instr_total * 100) if instr_total > 0 else 0
    )

    ws.append(
        [
            "Инструктажи",
            instr_total,
            instr_passed,
            instr_failed,
            f"{instr_success_rate:.1f}%",
        ]
    )

    # Итоговая строка
    total_total = test_total + instr_total
    total_passed = test_passed + instr_passed
    total_failed = test_failed + instr_failed
    total_success_rate = (
        (total_passed / total_total * 100) if total_total > 0 else 0
    )

    ws.append(
        [
            "Общий итог",
            total_total,
            total_passed,
            total_failed,
            f"{total_success_rate:.1f}%",
        ]
    )

    # Стили для данных
    for row in ws.iter_rows(min_row=6, max_row=8, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column in [2, 3, 4, 5]:
                cell.alignment = center_aligned

    # Подсветка итоговой строки
    for col in range(1, 6):
        cell = ws.cell(row=8, column=col)
        cell.fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        cell.font = Font(bold=True)

    # 3. Проблемные вопросы
    ws.append([])
    ws.append(["Самые проблемные вопросы (по количеству ошибок)"])
    ws.merge_cells("A10:E10")
    ws["A10"].font = Font(size=12, bold=True)
    ws["A10"].fill = PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    )
    ws["A10"].alignment = center_aligned

    headers = [
        "Вопрос",
        "Тест",
        "Количество ошибок",
        "ID вопроса",
        "Доля ошибок",
    ]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=11, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    for question in context["problematic_questions"]:
        error_rate = (
            (question["total_errors"] / test_total * 100)
            if test_total > 0
            else 0
        )
        ws.append(
            [
                question["question__name"],
                question.get("test_name", "-"),
                question["total_errors"],
                question["question_id"],
                f"{error_rate:.1f}%",
            ]
        )

    # 4. Сотрудники, требующие внимания
    ws.append([])
    ws.append(["Сотрудники с наибольшим количеством ошибок"])
    ws.merge_cells("A13:E13")
    ws["A13"].font = Font(size=12, bold=True)
    ws["A13"].fill = PatternFill(
        start_color="C00000", end_color="C00000", fill_type="solid"
    )
    ws["A13"].font = Font(color="FFFFFF")
    ws["A13"].alignment = center_aligned

    headers = [
        "Сотрудник",
        "Должность",
        "Проваленные тесты",
        "Проваленные инструктажи",
        "Всего провалов",
    ]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=14, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    for user in context["weak_users"]:
        total_fails = user["test_fails"] + user["instruction_fails"]
        ws.append(
            [
                user["user_name"],
                user.get("position", "-"),
                user["test_fails"],
                user["instruction_fails"],
                total_fails,
            ]
        )

        # Подсветка самых проблемных сотрудников
        if total_fails >= 3:
            for col in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )

    # Настройка ширины столбцов
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[column_letter]:
            try:
                value_length = len(str(cell.value)) if cell.value else 0
                if value_length > max_length:
                    max_length = value_length
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # Применяем стили ко всем данным
    for row in ws.iter_rows(
        min_row=6, max_row=ws.max_row, min_col=1, max_col=5
    ):
        for cell in row:
            cell.border = thin_border
            if cell.column in [3, 4, 5]:
                cell.alignment = center_aligned

    wb.save(response)
    return response

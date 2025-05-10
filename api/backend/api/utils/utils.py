import base64
from datetime import datetime
import json
import os

from Crypto.Cipher import AES
from dotenv import load_dotenv
from django.conf import settings
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from api.admin import dashboard_callback
from api.models import User


def is_face_already_registered(input_descriptor):
    """
    Проверяет, есть ли в системе пользователь с похожим дескриптором лица
    :param input_descriptor: numpy array с дескриптором лица
    :return: True если найден похожий пользователь, иначе False
    """
    for user in User.objects.exclude(face_descriptor__isnull=True):
        try:
            encrypted_data = json.loads(user.face_descriptor)
            stored_descriptor = np.array(
                decrypt_descriptor(
                    encrypted_data,
                    AES_STORAGE_KEY
                ), dtype=np.float32
            )

            distance = np.linalg.norm(input_descriptor - stored_descriptor)
            if distance < settings.FACE_MATCH_THRESHOLD:
                return True
        except Exception as e:
            print(f"Ошибка обработки дескриптора лица пользователя "
                  f"{user.id}: {str(e)}")
            continue
    return False


def decrypt_descriptor(encrypted_data, key):
    """Дешифрует дескриптор лица"""
    try:
        iv = base64.b64decode(encrypted_data['iv'])
        ciphertext = base64.b64decode(encrypted_data['ciphertext'])
        tag = base64.b64decode(encrypted_data['tag'])

        cipher = AES.new(key, AES.MODE_GCM, nonce=iv)
        decrypted = cipher.decrypt_and_verify(ciphertext, tag)
        return json.loads(decrypted.decode('utf-8'))
    except Exception as e:
        raise ValueError(f"Ошибка дешифрования: {str(e)}")


def encrypt_descriptor(descriptor, key):
    """Шифрует дескриптор лица для хранения"""
    try:
        descriptor_json = json.dumps(
            descriptor.tolist()
            if isinstance(descriptor, np.ndarray)
            else descriptor
        )

        cipher = AES.new(key, AES.MODE_GCM)
        ciphertext, tag = cipher.encrypt_and_digest(
            descriptor_json.encode('utf-8')
        )

        return {
            'iv': base64.b64encode(cipher.nonce).decode('utf-8'),
            'ciphertext': base64.b64encode(ciphertext).decode('utf-8'),
            'tag': base64.b64encode(tag).decode('utf-8')
        }
    except Exception as e:
        raise ValueError(f"Ошибка шифрования: {str(e)}")


@staff_member_required
def export_to_excel(request):
    # Получаем данные
    context = {}
    dashboard_callback(request, context)

    # Создаем Excel файл
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = (
        f'attachment; filename="full_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Полный отчёт"

    # Стили
    header_fill = PatternFill(
        start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    center_aligned = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 1. Шапка отчёта
    ws.merge_cells("A1:E1")
    ws["A1"] = "Полный отчёт по обучению сотрудников"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center_aligned

    ws["A2"] = "Дата формирования отчёта:"
    ws["B2"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["B2"].font = Font(bold=True)

    # 2. Основная статистика
    ws.append([])
    ws.append(["Детальная статистика по тестам и инструктажам"])
    ws.merge_cells("A4:E4")
    ws["A4"].font = Font(size=12, bold=True)
    ws["A4"].fill = PatternFill(
        start_color="70AD47", end_color="70AD47", fill_type="solid"
    )
    ws["A4"].font = Font(color="FFFFFF")
    ws["A4"].alignment = center_aligned

    # Заголовки таблицы
    headers = [
        "Тип проверки",
        "Всего попыток",
        "Успешно пройдено",
        "Не пройдено",
        "Процент успеха",
    ]
    ws.append(headers)

    # Стили для заголовков
    for col in range(1, 6):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    # Данные тестов
    test_total = context["test_stats"]["total"]
    test_passed = context["test_stats"]["passed"]
    test_failed = context["test_stats"]["failed"]
    test_success_rate = (
        (test_passed / test_total * 100) if test_total > 0 else 0
    )

    ws.append(
        [
            "Тесты",
            test_total,
            test_passed,
            test_failed,
            f"{test_success_rate:.1f}%",
        ]
    )

    # Данные инструктажей
    instr_total = context["instruction_stats"]["total"]
    instr_passed = context["instruction_stats"]["passed"]
    instr_failed = context["instruction_stats"]["failed"]
    instr_success_rate = (
        (instr_passed / instr_total * 100) if instr_total > 0 else 0
    )

    ws.append(
        [
            "Инструктажи",
            instr_total,
            instr_passed,
            instr_failed,
            f"{instr_success_rate:.1f}%",
        ]
    )

    # Итоговая строка
    total_total = test_total + instr_total
    total_passed = test_passed + instr_passed
    total_failed = test_failed + instr_failed
    total_success_rate = (
        (total_passed / total_total * 100) if total_total > 0 else 0
    )

    ws.append(
        [
            "Общий итог",
            total_total,
            total_passed,
            total_failed,
            f"{total_success_rate:.1f}%",
        ]
    )

    # Стили для данных
    for row in ws.iter_rows(min_row=6, max_row=8, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column in [2, 3, 4, 5]:
                cell.alignment = center_aligned

    # Подсветка итоговой строки
    for col in range(1, 6):
        cell = ws.cell(row=8, column=col)
        cell.fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        cell.font = Font(bold=True)

    # 3. Проблемные вопросы
    ws.append([])
    ws.append(["Самые проблемные вопросы (по количеству ошибок)"])
    ws.merge_cells("A10:E10")
    ws["A10"].font = Font(size=12, bold=True)
    ws["A10"].fill = PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    )
    ws["A10"].alignment = center_aligned

    headers = [
        "Вопрос",
        "Тест",
        "Количество ошибок",
        "ID вопроса",
        "Доля ошибок",
    ]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=11, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    for question in context["problematic_questions"]:
        error_rate = (
            (question["total_errors"] / test_total * 100)
            if test_total > 0
            else 0
        )
        ws.append(
            [
                question["question__name"],
                question.get("test_name", "-"),
                question["total_errors"],
                question["question_id"],
                f"{error_rate:.1f}%",
            ]
        )

    # 4. Сотрудники, требующие внимания
    ws.append([])
    ws.append(["Сотрудники с наибольшим количеством ошибок"])
    ws.merge_cells("A13:E13")
    ws["A13"].font = Font(size=12, bold=True)
    ws["A13"].fill = PatternFill(
        start_color="C00000", end_color="C00000", fill_type="solid"
    )
    ws["A13"].font = Font(color="FFFFFF")
    ws["A13"].alignment = center_aligned

    headers = [
        "Сотрудник",
        "Должность",
        "Проваленные тесты",
        "Проваленные инструктажи",
        "Всего провалов",
    ]
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=14, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_aligned
        cell.border = thin_border

    for user in context["weak_users"]:
        total_fails = user["test_fails"] + user["instruction_fails"]
        ws.append(
            [
                user["user_name"],
                user.get("position", "-"),
                user["test_fails"],
                user["instruction_fails"],
                total_fails,
            ]
        )

        # Подсветка самых проблемных сотрудников
        if total_fails >= 3:
            for col in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )

    # Настройка ширины столбцов
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[column_letter]:
            try:
                value_length = len(str(cell.value)) if cell.value else 0
                if value_length > max_length:
                    max_length = value_length
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # Применяем стили ко всем данным
    for row in ws.iter_rows(
        min_row=6, max_row=ws.max_row, min_col=1, max_col=5
    ):
        for cell in row:
            cell.border = thin_border
            if cell.column in [3, 4, 5]:
                cell.alignment = center_aligned

    wb.save(response)
    return response
